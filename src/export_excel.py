# export_excel.py

import os
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.formatting.rule import ColorScaleRule

EXPORT_DIR = "exports"
os.makedirs(EXPORT_DIR, exist_ok=True)

def export_to_excel(dataframes_dict, filename):
    """
    Exports multiple dataframes to an Excel file with formatting.

    Args:
        dataframes_dict (dict): Dictionary of {sheet_name: dataframe}
        filename (str): Output Excel filename
    """
    # Convert timezone-aware datetime columns to naive (Excel doesn't support tz-aware)
    for df in dataframes_dict.values():
        for col in df.select_dtypes(include=['datetimetz']):
            df[col] = df[col].dt.tz_localize(None)

    filepath = os.path.join(EXPORT_DIR, filename)

    # Write all dataframes to Excel
    with pd.ExcelWriter(filepath, engine="openpyxl") as writer:
        for sheet_name, df in dataframes_dict.items():
            df.to_excel(writer, sheet_name=sheet_name, index=False)
    
    # Load workbook to apply formatting
    wb = load_workbook(filepath)
    
    for sheet_name, df in dataframes_dict.items():
        ws = wb[sheet_name]

        # Freeze top row and first column
        ws.freeze_panes = "B2"

        # Apply auto filter
        ws.auto_filter.ref = ws.dimensions

        # Apply gradient conditional formatting to numeric columns
        for col_idx, col in enumerate(df.columns, start=1):
            if pd.api.types.is_numeric_dtype(df[col]):
                col_letter = ws.cell(row=1, column=col_idx).column_letter
                rule = ColorScaleRule(
                    start_type="min", start_color="FFAA0000",
                    mid_type="percentile", mid_value=50, mid_color="FFFFFF00",
                    end_type="max", end_color="FF00AA00"
                )
                ws.conditional_formatting.add(f"{col_letter}2:{col_letter}{ws.max_row}", rule)
        
        # Auto-adjust column widths
        for col_idx, col in enumerate(df.columns, start=1):
            max_length = max(
                df[col].astype(str).map(len).max(),
                len(str(col))
            ) + 2  # add a little padding
            ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = max_length
    
    # Save workbook
    wb.save(filepath)

    # Print console message
    total_rows = sum(len(df) for df in dataframes_dict.values())
    print(f"Created file {filename}, {len(dataframes_dict)} sheets, {total_rows} rows")
